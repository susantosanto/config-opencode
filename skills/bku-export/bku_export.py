#!/usr/bin/env python3
"""
BKU Export Tool - Export Buku Kas Umum dari ARKAS ke Excel.

Mendukung export per bulan, per tahun, dengan format debit/kredit/saldo.

Usage:
  python bku_export.py                          # Export tahun berjalan
  python bku_export.py --bulan 3               # Export Maret tahun ini
  python bku_export.py --bulan 3 --tahun 2025  # Export Maret 2025
  python bku_export.py --tahun 2025            # Export full tahun 2025
  python bku_export.py --bulan 1-3             # Export Jan-Maret tahun ini
  python bku_export.py --list                  # Lihat bulan dengan data
  python bku_export.py --output custom.xlsx    # Kustom nama file output
  python bku_export.py --format rekapitulasi   # Format rekapitulasi per bulan
"""

import json
import os
import sys
import argparse
from datetime import datetime, date

# ─── Konfigurasi ──────────────────────────────────────
CONFIG_PATH = os.path.expanduser(r"~\.config\opencode\arkas_config.json")
EXPORT_DIR = os.path.expanduser(r"~\Documents\arkas_analysis\exports")
BULAN_NAMES = {
    1: "Januari", 2: "Februari", 3: "Maret", 4: "April",
    5: "Mei", 6: "Juni", 7: "Juli", 8: "Agustus",
    9: "September", 10: "Oktober", 11: "November", 12: "Desember"
}

# Klasifikasi tipe transaksi berdasarkan id_ref_bku
DEBIT_TYPES   = {1, 2, 6, 8, 9, 10, 23, 25, 26, 28, 29, 30}
KREDIT_TYPES  = {3, 4, 5, 7, 11, 12, 13, 14, 15, 24, 27, 31, 32, 33, 34, 35}


# ─── Koneksi Database ─────────────────────────────────
def load_config():
    """Load konfigurasi ARKAS."""
    if not os.path.exists(CONFIG_PATH):
        print(f"❌ Config tidak ditemukan: {CONFIG_PATH}")
        sys.exit(1)

    with open(CONFIG_PATH, "r") as f:
        cfg = json.load(f)

    return cfg["arkas"]


def connect_db(cfg):
    """Koneksi ke database ARKAS SQLCipher."""
    try:
        import sqlcipher3 as sqlite
    except ImportError:
        print("❌ sqlcipher3 tidak terinstall.")
        print("   Install: pip install sqlcipher3-binary")
        sys.exit(1)

    if not os.path.exists(cfg["db_path"]):
        print(f"❌ Database tidak ditemukan: {cfg['db_path']}")
        sys.exit(1)

    db = sqlite.connect(cfg["db_path"])
    db.execute(f"PRAGMA key = '{cfg['key']}'")
    db.execute(f"PRAGMA cipher_compatibility = {cfg['cipher_compatibility']}")
    return db


# ─── Ambil Data ────────────────────────────────────────
def get_ref_bku_map(db):
    """Ambil mapping id_ref_bku ke nama untuk informasi."""
    try:
        cursor = db.execute("SELECT id_ref_bku, nama FROM ref_bku")
        return {r[0]: r[1] for r in cursor.fetchall()}
    except Exception:
        return {}


def get_transactions(db, tahun, bulan_awal=None, bulan_akhir=None):
    """
    Ambil transaksi kas_umum berdasarkan periode.
    
    Args:
        db: Koneksi database
        tahun: Tahun (int)
        bulan_awal: Bulan awal (int, 1-12) atau None untuk semua
        bulan_akhir: Bulan akhir (int, 1-12), None = sama dengan bulan_awal
    
    Returns:
        List of dict transaksi
    """
    if bulan_awal and not bulan_akhir:
        bulan_akhir = bulan_awal

    if bulan_awal and bulan_akhir:
        tgl_awal = f"{tahun}-{bulan_awal:02d}-01"
        if bulan_akhir == 12:
            tgl_akhir = f"{tahun + 1}-01-01"
        else:
            tgl_akhir = f"{tahun}-{bulan_akhir + 1:02d}-01"

        sql = """
            SELECT tanggal_transaksi, no_bukti, uraian, saldo, id_ref_bku
            FROM kas_umum
            WHERE tanggal_transaksi >= ? AND tanggal_transaksi < ?
            AND soft_delete = 0
            ORDER BY tanggal_transaksi ASC, create_date ASC
        """
        params = (tgl_awal, tgl_akhir)
    else:
        sql = """
            SELECT tanggal_transaksi, no_bukti, uraian, saldo, id_ref_bku
            FROM kas_umum
            WHERE tanggal_transaksi LIKE ?
            AND soft_delete = 0
            ORDER BY tanggal_transaksi ASC, create_date ASC
        """
        params = (f"{tahun}-%",)

    cursor = db.execute(sql, params)
    rows = cursor.fetchall()
    return rows


def classify_transaction(bku_type, amount):
    """Klasifikasikan transaksi sebagai debit atau kredit."""
    if bku_type in DEBIT_TYPES:
        return amount, 0
    elif bku_type in KREDIT_TYPES:
        return 0, amount
    else:
        return 0, 0


def process_transactions(rows, ref_bku_map=None):
    """Proses transaksi menjadi format dengan debit/kredit/saldo."""
    data = []
    current_balance = 0

    for row in rows:
        tanggal, no_bukti, uraian, amount, bku_type = row
        debit, kredit = classify_transaction(bku_type, amount)
        current_balance += debit - kredit

        jenis = "DEBET" if debit > 0 else ("KREDIT" if kredit > 0 else "-")
        nama_ref = ref_bku_map.get(bku_type, f"Tipe {bku_type}") if ref_bku_map else f"Tipe {bku_type}"

        data.append({
            "Tanggal": tanggal,
            "No. Bukti": no_bukti or "-",
            "Uraian": uraian,
            "Jenis": jenis,
            "Ref BKU": nama_ref,
            "Debit": debit,
            "Kredit": kredit,
            "Saldo": current_balance
        })

    return data


# ─── Export ────────────────────────────────────────────
def export_to_excel(data, output_path, sekolah, periode_label):
    """Export data BKU ke file Excel dengan formatting."""
    try:
        import pandas as pd
        from openpyxl import load_workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
    except ImportError:
        print("❌ pandas/openpyxl tidak terinstall.")
        print("   Install: pip install pandas openpyxl")
        return False

    if not data:
        print("  ⚠️  Tidak ada data untuk diexport.")
        return False

    df = pd.DataFrame(data)

    # Format kolom currency
    currency_cols = ["Debit", "Kredit", "Saldo"]
    for col in currency_cols:
        df[col] = df[col].apply(lambda x: f"Rp {x:,.0f}")

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    df.to_excel(output_path, index=False, sheet_name="BKU")

    # Styling Excel
    wb = load_workbook(output_path)
    ws = wb.active

    # Styling: header
    header_fill = PatternFill(start_color="17294D", end_color="17294D", fill_type="solid")
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    # Auto-width
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 3, 50)

    # Freeze pane
    ws.freeze_panes = "A2"

    wb.save(output_path)

    # Ringkasan
    total_debit = sum(d["Debit"] for d in data)
    total_kredit = sum(d["Kredit"] for d in data)
    saldo_akhir = data[-1]["Saldo"] if data else 0

    print(f"\n{'='*60}")
    print(f"  ✅ EXPORT BERHASIL")
    print(f"{'='*60}")
    print(f"  📁 File : {output_path}")
    print(f"  🏫 Sekolah: {sekolah}")
    print(f"  📅 Periode: {periode_label}")
    print(f"  📊 Transaksi: {len(data)} baris")
    print(f"  💰 Total Debit : Rp {total_debit:,.0f}")
    print(f"  💸 Total Kredit: Rp {total_kredit:,.0f}")
    print(f"  💵 Saldo Akhir: Rp {saldo_akhir:,.0f}")
    print(f"{'='*60}")

    return True


def export_rekapitulasi(db, cfg, tahun):
    """Export rekapitulasi BKU per bulan (format ringkas)."""
    try:
        import pandas as pd
    except ImportError:
        print("❌ pandas tidak terinstall.")
        return

    data_bulanan = []
    for bulan in range(1, 13):
        rows = get_transactions(db, tahun, bulan, bulan)
        if not rows:
            continue

        total_debit = 0
        total_kredit = 0
        for r in rows:
            d, k = classify_transaction(r[4], r[3])
            total_debit += d
            total_kredit += k

        data_bulanan.append({
            "Bulan": BULAN_NAMES.get(bulan, f"Bulan {bulan}"),
            "Bulan Ke": bulan,
            "Transaksi": len(rows),
            "Total Debit": total_debit,
            "Total Kredit": total_kredit,
            "Selisih": total_debit - total_kredit
        })

    if not data_bulanan:
        print(f"  ⚠️  Tidak ada data BKU tahun {tahun}")
        return

    df = pd.DataFrame(data_bulanan)

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    path = os.path.join(EXPORT_DIR, f"BKU_Rekapitulasi_{tahun}_{ts}.xlsx")
    os.makedirs(EXPORT_DIR, exist_ok=True)
    df.to_excel(path, index=False)

    print(f"\n{'='*60}")
    print(f"  ✅ REKAPITULASI BKU {tahun}")
    print(f"{'='*60}")
    for _, r in df.iterrows():
        print(f"  {r['Bulan']:12s} | {r['Transaksi']:3d} tx | "
              f"Debit: Rp {r['Total Debit']:>10,.0f} | "
              f"Kredit: Rp {r['Total Kredit']:>10,.0f}")
    print(f"{'='*60}")
    print(f"  📁 File: {path}")


def list_available_periods(db):
    """Tampilkan bulan-bulan yang memiliki data BKU."""
    cursor = db.execute("""
        SELECT strftime('%Y', tanggal_transaksi) as tahun,
               CAST(strftime('%m', tanggal_transaksi) AS INTEGER) as bulan,
               COUNT(*) as jumlah
        FROM kas_umum
        WHERE soft_delete = 0
        GROUP BY tahun, bulan
        ORDER BY tahun DESC, bulan DESC
    """)
    periods = cursor.fetchall()

    print(f"\n{'='*60}")
    print(f"  📅 PERIODE BKU TERSEDIA")
    print(f"{'='*60}")
    for p in periods:
        nama_bulan = BULAN_NAMES.get(p[1], f"Bulan {p[1]}")
        print(f"  {p[0]} - {nama_bulan:10s} ({p[2]} transaksi)")
    print(f"{'='*60}")


# ─── CLI Main ─────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(
        description="BKU Export - Export Buku Kas Umum ARKAS ke Excel",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Contoh:
  python bku_export.py                        # Tahun ini
  python bku_export.py --bulan 3             # Maret tahun ini
  python bku_export.py --bulan 3 --tahun 2025  # Maret 2025
  python bku_export.py --tahun 2025          # Full tahun 2025
  python bku_export.py --bulan 1-3           # Jan-Mar tahun ini
  python bku_export.py --list                # Periode tersedia
  python bku_export.py --format rekapitulasi # Ringkasan per bulan
  python bku_export.py --output laporan.xlsx # Nama file kustom
        """
    )

    parser.add_argument("--bulan", help="Bulan (1-12) atau range (1-3)")
    parser.add_argument("--tahun", type=int, help="Tahun (default: tahun berjalan)")
    parser.add_argument("--list", action="store_true", help="Lihat periode tersedia")
    parser.add_argument("--output", help="Path file output kustom")
    parser.add_argument("--format", choices=["detail", "rekapitulasi"], default="detail",
                        help="Format output (default: detail)")

    args = parser.parse_args()

    cfg = load_config()
    db = connect_db(cfg)

    sekolah = cfg.get("sekolah", "Unknown")
    tahun_sekarang = date.today().year

    print(f"\n  🏫 {sekolah}")
    print(f"  {'='*50}")

    # List periode
    if args.list:
        list_available_periods(db)
        db.close()
        return

    # Format rekapitulasi
    if args.format == "rekapitulasi":
        tahun = args.tahun or tahun_sekarang
        export_rekapitulasi(db, cfg, tahun)
        db.close()
        return

    # Parse bulan
    bulan_awal = None
    bulan_akhir = None
    if args.bulan:
        if "-" in args.bulan:
            parts = args.bulan.split("-")
            bulan_awal = int(parts[0])
            bulan_akhir = int(parts[1])
        else:
            bulan_awal = int(args.bulan)

    tahun = args.tahun or tahun_sekarang

    # Buat label periode
    if bulan_awal and bulan_akhir and bulan_awal != bulan_akhir:
        nama_awal = BULAN_NAMES.get(bulan_awal, str(bulan_awal))
        nama_akhir = BULAN_NAMES.get(bulan_akhir, str(bulan_akhir))
        periode_label = f"{nama_awal} - {nama_akhir} {tahun}"
    elif bulan_awal:
        nama_bulan = BULAN_NAMES.get(bulan_awal, str(bulan_awal))
        periode_label = f"{nama_bulan} {tahun}"
    else:
        periode_label = f"Tahun {tahun}"

    print(f"  📅 Periode: {periode_label}")

    # Ambil data
    rows = get_transactions(db, tahun, bulan_awal, bulan_akhir)
    if not rows:
        print(f"\n  ⚠️  Tidak ada transaksi BKU untuk {periode_label}")
        db.close()
        return

    ref_map = get_ref_bku_map(db)
    data = process_transactions(rows, ref_map)

    print(f"  📊 Transaksi: {len(data)} item")

    # Export
    if args.output:
        output_path = args.output
    else:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        if bulan_awal and bulan_akhir:
            fname = f"BKU_{BULAN_NAMES.get(bulan_awal, bulan_awal)}_{BULAN_NAMES.get(bulan_akhir, bulan_akhir)}_{tahun}_{ts}.xlsx"
        elif bulan_awal:
            fname = f"BKU_{BULAN_NAMES.get(bulan_awal, bulan_awal)}_{tahun}_{ts}.xlsx"
        else:
            fname = f"BKU_Tahunan_{tahun}_{ts}.xlsx"
        output_path = os.path.join(EXPORT_DIR, fname)

    export_to_excel(data, output_path, sekolah, periode_label)

    db.close()


if __name__ == "__main__":
    main()
